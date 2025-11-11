# %%
from difflib import get_close_matches
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment


dic_quinzena = {
    '1': 'Primeira Quinzena',
    '2': 'Segunda Quinzena'
}
dic_mes = {
    '01': 'Janeiro',
    '02': 'Fevereiro',
    '03': 'Março',
    '04': 'Abril',
    '05': 'Maio',
    '06': 'Junho',
    '07': 'Julho',
    '08': 'Agosto',
    '09': 'Setembro',
    '10': 'Outubro',
    '11': 'Novembro',
    '12': 'Dezembro'
}

def cria_quinzena(a):
    if a == 0:
        return 0
    else:
        quinzena = str(a)[0]
        mes = str(a)[1:]
        nome_quinzena = f'{dic_quinzena[quinzena]} de {dic_mes[mes]}'
        return nome_quinzena

def find_closest_sku(sku, sku_list):
    matches = get_close_matches(
        sku, sku_list, n=1, cutoff=0.20
    )  # Ajuste o cutoff conforme necessário
    return matches[0] if matches else None


st.set_page_config(layout="wide", page_icon='💵', page_title='Fechamento - Eddi Casa')
st.title("Fechamento - Mão de Obra")

url_valores = "https://docs.google.com/spreadsheets/d/1JkafGyVeOQjCvMmePSfrgW3rSrTs6bzcL01G1Spge4s/export?format=csv&gid=2104680401#gid=2104680401"

df_valores = pd.read_csv(url_valores)
df_valores["MO"] = df_valores["MO"].str.title()

url_inicio = "https://docs.google.com/spreadsheets/d/1cGeQrjvsnuj9K1S_uPrYwxDKnUoyHnQEvFJjttU4Pcw/"
url_fim = "gid=887283048#gid=887283048"
url = f"{url_inicio}export?format=csv&{url_fim}"

df = pd.read_csv(url)
df["MAXHOME"] = df["MAXHOME"].str.title()
df = df.drop(
    columns=[
        "Descrição do produto",
        "Destino",
        "% Completude",
        "Dias na MO",
        "Setor de finalização",
        "Preço unitário",
        "Concluído",
    ]
)

df["Observações"] = df["Observações"].fillna(0).astype(int)

# ---------------- FILTRO 1 - Quinzena -------------

col1, col2, col3, col4 = st.columns(4)

df['Nome Quinzena'] = df['Observações'].apply(cria_quinzena)

lista_quinzenas = df['Nome Quinzena'].unique().tolist()
lista_quinzenas.remove(0)
lista_organizada = lista_quinzenas[::-1]
filtro1 = col1.selectbox(label='Quinzena', options=lista_organizada)

filtro_quinzena = df["Nome Quinzena"] == filtro1

df = df[filtro_quinzena]

# ---------------- FILTRO 2 - Mao de obra -------------

mo_quinzena = df["MAXHOME"].unique().tolist()
mo_quinzena = sorted(mo_quinzena)
mo_selecionada = col2.selectbox(label="Mão de Obra", options=mo_quinzena)

filtro2 = df["MAXHOME"] == mo_selecionada

df = df[filtro2]

filtro3 = df_valores["MO"] == mo_selecionada
df_valores = df_valores[filtro3]

df["Sku_mapeado"] = df["Código (SKU) "].apply(
    lambda x: find_closest_sku(x, df_valores["SKU"])
)

df = df.merge(
    right=df_valores,
    left_on=["Sku_mapeado"],
    right_on=["SKU"],
    how="left",
    suffixes=["Producao", "Valores"],
)

df = df.drop(columns=["MAXHOME", "Observações","Nome Quinzena", "Sku_mapeado", "SKU", "MO"])
df["VALOR"] = df["VALOR"].str.replace(",", ".").astype(float)

df["Total"] = df["Quantidade"] * df["VALOR"]

quantidade_total = df["Quantidade"].sum()

total = df["Total"].sum()
col3.metric(label="Valor", value=f"R$ {total:.2f}")
col4.metric(label="Quantidade entregue", value=quantidade_total)

df.columns = ["Data", "Requisição", "SKU", "Quantidade", "Pedido", "Valor", "Total"]
ordem_colunas = ["Data", "Pedido", "Requisição", "SKU", "Quantidade", "Valor", "Total"]
df = df[ordem_colunas]
df1 = df
df = df.set_index(df["Data"]).drop(columns="Data")

columns_config = {
    "Valor": st.column_config.NumberColumn("Valor", format="R$ %.2f", ),
    "Total": st.column_config.NumberColumn("Total", format="R$ %.2f"),
}

if st.button('Clique aqui para baixar'):
    file = 'fechamento.xlsx'
    wb = load_workbook(file)
    ws = wb.active
    ws['c2'] = filtro1
    ws['c3'] = mo_selecionada
    novo_arquivo = f'{filtro1} - {mo_selecionada}.xlsx'
    wb.save(novo_arquivo)

    with pd.ExcelWriter(novo_arquivo, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer: 
        df1.to_excel(writer, startrow=5, startcol=0, index=False, header=False)

    file_path = novo_arquivo
    sheet_name = "Sheet1"  # Troque pelo nome da sua aba!
    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    # Descobrir o range existente
    max_row = ws.max_row
    max_col = ws.max_column
    start_row = 6
    for row in ws.iter_rows(min_row=start_row, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(file_path)

st.dataframe(df, column_config=columns_config)

