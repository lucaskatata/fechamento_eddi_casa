# %%
import pandas as pd
from openpyxl import load_workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment

mo = 'kaizen'
quinzena = 'primeira quinzena de novembro'

df = pd.read_csv('kaizen.csv')

file = 'fechamento.xlsx'
wb = load_workbook(file)
ws = wb.active
ws['c2'] = quinzena
ws['c3'] = mo
novo_arquivo = 'kaizen.xlsx'
wb.save(novo_arquivo)

with pd.ExcelWriter('kaizen.xlsx', engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer: 
    df.to_excel(writer, startrow=5, startcol=0, index=False, header=False)


file_path = "kaizen.xlsx"
sheet_name = "Sheet1"  # Troque pelo nome da sua aba!

wb = load_workbook(file_path)
ws = wb[sheet_name]

# Descobrir o range existente
max_row = ws.max_row
max_col = ws.max_column
start_row = 6

for row in ws.iter_rows(min_row=start_row, max_row=max_row, min_col=1, max_col=max_col):
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center")

wb.save(file_path)
print("✅ Formatação aplicada: centralizado!")